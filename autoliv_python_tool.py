import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def list_folders_in_directory(directory):
    try:
        items = os.listdir(directory)
        folders = [item for item in items if os.path.isdir(os.path.join(directory, item)) and item != '.ipynb_checkpoints']
        print("Your folder contains:")

        for idx, folder_name in enumerate(folders, start=1):  
            print(f'{idx} : {folder_name}')
        print('\n')
        
        while True:
            try:
                selected_folder_no = int(input("Enter index of required folder: "))
                if 1 <= selected_folder_no <= len(folders):
                    selected_folder = folders[selected_folder_no - 1]
                    break
                else:
                    print(f'Invalid number. Please enter a number between 1 and {len(folders)}.')
            except ValueError:
                print("Invalid input. Please enter a numeric value.")

        print("\nSelected folder is:", selected_folder)
        selected_folder_path = os.path.join(directory, selected_folder)
        contents = os.listdir(selected_folder_path)

        # Filter to get only Excel files
        excel_files = [file for file in contents if file.endswith('.xlsx') or file.endswith('.xls')]
        
        if not excel_files:
            print(f"The folder '{selected_folder}' does not contain any Excel files.")
            return [], None, None
        
        print("Excel files in the selected folder:")
        for idx, file in enumerate(excel_files, start=1):
            print(f"{idx} : {file}")
        
        while True:
            try:
                selected_file_no = int(input("Enter index of the Excel file to choose: "))
                if 1 <= selected_file_no <= len(excel_files):
                    selected_file = excel_files[selected_file_no - 1]
                    break
                else:
                    print(f'Invalid number. Please enter a number between 1 and {len(excel_files)}.')
            except ValueError:
                print("Invalid input. Please enter a numeric value.")
        
        print(f"\nSelected file is: {selected_file}")
        selected_file_path = os.path.join(selected_folder_path, selected_file)

        return excel_files, selected_file_path, selected_folder

    except FileNotFoundError:
        print(f"The directory {directory} does not exist.")
        return [], None, None
    except PermissionError:
        print(f"You do not have permission to access {directory}.")
        return [], None, None

def get_info(file_adr):
    try:
        wb = load_workbook(file_adr)
    except FileNotFoundError:
        print('File not found. Please enter a valid file location.')
        return []
    except Exception as e:
        print(f'Error loading file: {e}')
        return []

    sheets = wb.sheetnames
    valid_sheets = []

    # Display all sheets and skip those with charts
    print("The file contains the following sheets:")
    for idx, sheet in enumerate(sheets, start=1):
        ws = wb[sheet]
        
        # Check if the sheet contains charts
        if not ws._charts:  # If no charts are present in the sheet
            valid_sheets.append(sheet)
            print(f'{idx} : {sheet}')
        else:
            print(f'Skipping sheet "{sheet}" as it contains charts.')

    # Let the user select the desired sheets
    if valid_sheets:
        print('Enter the index of the sheets to be selected (e.g., 1,2,3 )')
        while True:
            try:
                selected_sheet_indices = list(map(int, input().replace(',', ' ').split()))
                valid_selected_sheets = [valid_sheets[i - 1] for i in selected_sheet_indices if 1 <= i <= len(valid_sheets)]
                if valid_selected_sheets:
                    break
                else:
                    print(f"Error: Input out of range! Please select between 1 and {len(valid_sheets)}.")
            except ValueError:
                print("Invalid input. Please enter numeric values separated by commas.")

        return valid_selected_sheets
    else:
        print("No valid sheets available for selection.")
        return []


def read_colNrow(file_adr, sheets):
    wb = load_workbook(file_adr)
    ws = wb[sheets[0]]
    headings = []

    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column, values_only=True):
        headings.append(row)
    return list(headings[0])

def extract_columns(rows):
    last_idx = len(rows)
    print("\nThe columns are:")
    for idx, value in enumerate(rows, start=1):
        print(idx, ":", value)
    
    columns = []
    multipliers = {}
    print('Enter index of columns to be selected (e.g., 1,2,3 )')

    while True:
        try:
            column = list(map(int, input().replace(',', ' ').split()))
            state = all(1 <= cols <= last_idx for cols in column)
            if state:
                break
            else:
                print('Error: Input out of range! Try again.')
        except ValueError:
            print("Invalid input. Please enter numeric values separated by commas.")

    for col in column:
        # Loop until valid input is provided for applying the multiplier
        while True:
            apply_multiplier = input(f"Do you want to multiply any constant value to the column '{rows[col-1]}'? (y/n): ").strip().lower()
            if apply_multiplier == 'y':
                while True:
                    try:
                        constant = float(input(f"What value of constant do you want to multiply to the column '{rows[col-1]}'?: "))
                        multipliers[col] = constant
                        break
                    except ValueError:
                        print("Invalid input. Please enter a numeric value.")
                break  # Exit the loop after valid multiplier input
            elif apply_multiplier == 'n':
                multipliers[col] = 1  # No multiplication, multiplier is 1
                break  # Exit the loop for this column
            else:
                print("Invalid input. Please enter 'y' or 'n'.")  # Invalid input case

    return column, multipliers


def read_data_from_colLst(path, sheet, columns_idx_list, multipliers):
    wb = load_workbook(path)
    data = []
    column_titles = []
    for i in range(len(sheet)):
        ws = wb[sheet[i]]
        for j in columns_idx_list:
            column_data = []
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=j, max_col=j, values_only=True):
                try:
                    value = float(row[0])
                    value *= multipliers[j]  # Apply the multiplier
                    column_data.append(value)
                except:
                    column_data.append(row[0])
            data.append(column_data)
            column_titles.append(ws.cell(row=1, column=j).value)  # Get the title of the column
    
    return data, column_titles

def create_excel_file():
    print('\n')
    filename = input('Enter new Excel file name: ')
    filename = filename + ".xlsx"
    desktop_path = os.path.join(directory, filename)
   
    wb = Workbook()
    wb.save(desktop_path)
    return desktop_path
from math import ceil

def write_into(file_adr, sheet_name, data, column_titles, file_name, num_sheets): 
    wb = load_workbook(file_adr)
    new_sheet = wb.create_sheet(title=sheet_name)
    ws = wb[sheet_name]

    # Write the file name in the first row
    fi=os.path.splitext(file_name)[0]
    ws['A1'] = fi

    sample_heading_counter = 1  # To track sample heading for each sheet

    # Replace specific column titles and add sample headings for the first column of each worksheet
    for idx, title in enumerate(column_titles):
        if title == "Longitudinal Engineering Strain [%]" or title == "Gage Lenght Longitudinal Engineering Strain [%]":
            column_titles[idx] = "Engineering strain[%]"
        elif title in [
            "Engineering stress [Pa]",
            "True stress CV [Pa]",
            "True stress HTS [Pa]"
        ]:
            column_titles[idx] = title.replace("[Pa]", "[MPa]")

    # Write the sample heading only for the first column from each worksheet
    col_counter = 0  # Tracks which sheet's data is being written
    for i in range(num_sheets):
        sample_heading = f"sample{sample_heading_counter}"
        ws[f'{get_column_letter(col_counter+1)}2'] = sample_heading  # Adding sample heading in the second row
        sample_heading_counter += 1
        col_counter += len(data) // num_sheets  # Move to the next sheet's columns after filling current sheet's data

    # Write the column titles starting from the third row
    for idx, title in enumerate(column_titles, start=1):
        ws[f'{get_column_letter(idx)}3'] = title

    # Write the data starting from the fourth row
    for col_idx, col_data in enumerate(data):
        for row_idx, value in enumerate(col_data):
            try:
                ws[f'{get_column_letter(col_idx+1)}{row_idx+4}'] = float(value)
            except ValueError:
                ws[f'{get_column_letter(col_idx+1)}{row_idx+4}'] = value

    wb.save(file_adr)

############################################################################################
new_F_data = []
data_none = [None]
directory = input('Enter file location:').replace('\\','\\\\').replace('"','')
file_created = False

while True:
    contents, path, selected_folder = list_folders_in_directory(directory)
    sheets = get_info(path)
    to_choose_heading = read_colNrow(path, sheets)
    columns_idx_list, multipliers = extract_columns(to_choose_heading)
    data, column_titles = read_data_from_colLst(path, sheets, columns_idx_list, multipliers)
    if not file_created:
        new_xl_add = create_excel_file()
        file_created = True
    sheet_name = input('Enter sheet name: ')
    write_into(new_xl_add, sheet_name, data, column_titles, os.path.basename(new_xl_add), len(sheets))
    exit_key = input('Enter "E" to exit or any other key to select columns from a different file: ').strip().lower()
    if exit_key == 'e':
        print('\nFile created successfully at:', new_xl_add, '\nTHANK YOU')
        break
    new_F_data = []
